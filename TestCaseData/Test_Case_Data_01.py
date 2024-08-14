import openpyxl
class TestDataOnce:
    DataSet = [{"Search_Item": "cu"}, {"Search_Item": "c"}, {"Search_Item": "p"}, {"Search_Item": "or"}]
    @staticmethod  # To Avoid Obj Creation and need not have self method
    def TestDataExcel(TestCaseName):
        book = openpyxl.load_workbook("C:\\Users\\Vikneshwaraj\\Documents\\PythonExcelData.xlsx")
        sheet = book.active
        Dict = {}
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == "Testcase_01":
                for j in range(2, sheet.max_column + 1):
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

        return [Dict]